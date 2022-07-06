from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from utils_.config import CUSTOMERS, SMALL_CUSTOMERS
from tqdm import tqdm

brightpurple1 = PatternFill(start_color='E8DAEF',
                            end_color='E8DAEF',
                            fill_type='solid')

brightgreen1 = PatternFill(start_color='D5F5E3',
                           end_color='D5F5E3',
                           fill_type='solid')

brightgreen2 = PatternFill(start_color='ABEBC6',
                           end_color='ABEBC6',
                           fill_type='solid')

brightyellow = PatternFill(start_color='F9E79F',
                           end_color='F9E79F',
                           fill_type='solid')

brightorange = PatternFill(start_color='F5CBA7',
                           end_color='F5CBA7',
                           fill_type='solid')

brightpurple2 = PatternFill(start_color='D2B4DE',
                            end_color='D2B4DE',
                            fill_type='solid')

brightblue = PatternFill(start_color='AED6F1',
                         end_color='AED6F1',
                         fill_type='solid')

lightgrey = PatternFill(start_color='EAEDED',
                        end_color='EAEDED',
                        fill_type='solid')

darkgrey = PatternFill(start_color='AEB6BF',
                       end_color='AEB6BF',
                       fill_type='solid')

empasizing_yellow = PatternFill(start_color='FFF975',
                                end_color='FFF975',
                                fill_type='solid')

thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='566573'),
    right=Side(border_style=BORDER_THIN, color='566573'),
    top=Side(border_style=BORDER_THIN, color='566573'),
    bottom=Side(border_style=BORDER_THIN, color='566573')
)


def fit_column_width(sheet):
    column_widths = []
    for row in sheet:
        for i, cell in enumerate(row):
            try:
                number = int(cell.value)
            except (ValueError, TypeError):
                length = len(str(cell.value)) * 2 if cell.value is not None else 0
            else:
                length = len(str(cell.value)) * 1.4

            if len(column_widths) > i:
                if length > column_widths[i]:
                    column_widths[i] = length
            else:
                column_widths += [length]

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = int(column_width)


def get_col_from_index(col, add):
    return get_column_letter(column_index_from_string(col) + add)


def formatting(xfile, df, sheetname, raw=False):
    sheet = xfile[sheetname]
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 95
    # fit_column_width(sheet)
    df.reset_index(inplace=True)
    sheet.column_dimensions["A"].width = 2

    row_anchor = 2
    col_anchor = "B"
    region_len = len(df)
    region_col_len = len(df.columns)-1

    for row in sheet:
        for col, cell in enumerate(row):
            if cell.value is not None:
                cell.border = thin_border
            if cell.value in df.columns.tolist():
                row[col].fill = brightgreen1
                if cell.value in ['차종', '코드', '코드.1', '작성자', '담당자', '현재결재자', '사업장', '담당자_', '담당공장']:
                    sheet.column_dimensions[get_column_letter(cell.column)].width = 8
                if cell.value in ['품번', '발생일자', '등록일자', '해외결재일', '국내접수일', '국내결재일', '제기일자', '제기부서',
                                  '검사일자', '통보일자', '접수일자', '결재일자', '포장장', '대책등록일', '대책확인일', '개선일',
                                  '불량유형1', '불량유형2', '불량유형3', 'P/T기종', 'P/T_Model', '고객사', 'KDQR_Type',
                                  'PPR_확인일자', '발생원인(대)', '발생원인(중)', '발생원인(소)', '대책등록일', '대책확인일',
                                  '대책확인일(해외)', '개선일', '공장', '불량원인(대)', '불량원인(중)', '불량원인(소)', '귀책구분',
                                  '불량유형(대)', '불량유형(중)', '불량유형(소)', 'KDQR_No', 'PPR_확인_일자', '해외_담당접수소요일',
                                  '국내_담당결재소요일', '담당고객사']:
                    sheet.column_dimensions[get_column_letter(cell.column)].width = 18
                if cell.value in ['부품협력사', '귀책처', '제목', '품명', '최종조립처', '긴급도']:
                    sheet.column_dimensions[get_column_letter(cell.column)].width = 35
            if not raw:
                if type(cell.value) == int:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.0'

    if raw:
            for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, region_col_len - 1)}{row_anchor + region_len}']:
                for col, cell in enumerate(row):
                    cell.border = thin_border
