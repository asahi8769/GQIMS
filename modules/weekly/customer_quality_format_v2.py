import openpyxl
from openpyxl.utils import get_column_letter
from modules.weekly.weekly_config import *
from utils_.config import CUSTOMERS, PERSON_IN_CHARGE
from datetime import datetime
from modules.weekly.pipelines.raw_data import get_raw_ovs, get_raw_dqy
from modules.weekly.customer_barplots_v2 import CustomerBarplots
from modules.weekly.customer_donutplots_v2 import CustomerDonutPlots
from modules.weekly.customer_delayed_v2 import CustomerDelayedPlots
from modules.weekly.pipelines.processed_data import OvsProcessedData
import shutil
from utils_.functions import make_dir
from modules.db.db_weekly_notes import WeeklyNotesModel
import os


class ExcelFormatSetter:

    def __init__(self, r_weeks=1, reprocess=REPROCESS, reload_charts=RELOAD_CHARTS, pre_noti=PRE_NOTI, monday_to_saturday=MONDAY_TO_SATURDAY,
                 regular_month=REGULAR_MONTH, yyyymm=YYYYMM, weekly_customer=WEEKLY_CUSTOMER, weekly_type=WEEKLY_TYPE,
                 n_weeks=N_WEEKS, n_ranks=N_RANKS):

        self.reload = reload_charts
        self.pre_noti = pre_noti
        self.monday_to_saturday = monday_to_saturday
        self.regular_month = regular_month
        self.yyyymm = yyyymm if regular_month else int(datetime.today().strftime("%Y%m"))
        self.weekly_customer = weekly_customer
        self.weekly_type = weekly_type
        self.n_weeks = n_weeks
        self.n_ranks = n_ranks

        self.raw_ovs_df = get_raw_ovs(pre_noti=self.pre_noti, regular_month=self.regular_month, yyyymm=self.yyyymm,
                                      weekly_customer=self.weekly_customer, weekly_type=self.weekly_type)
        self.raw_dqy_df = get_raw_dqy(regular_month=self.regular_month, yyyymm=self.yyyymm, weekly_customer=self.weekly_customer)

        self.data_obj = OvsProcessedData(self.raw_ovs_df, self.raw_dqy_df, pre_noti=self.pre_noti, monday_to_saturday=self.monday_to_saturday,
                                         regular_month=self.regular_month, yyyymm=self.yyyymm, weekly_customer=self.weekly_customer,
                                         weekly_type=self.weekly_type, n_weeks=self.n_weeks, n_ranks=self.n_ranks)

        self.df_weekly = None
        self.customer_sequence = W_CUSTOMERS if self.weekly_customer else CUSTOMERS
        self.df_dict, self.week_start, self.week_end = self.data_obj.weekly_overseas_rank(weeks=r_weeks)

        self.template_dir = "templates\customer_template.xlsx"
        if reprocess:
            self.processed_dir = "templates\customer_processed.xlsx"
            self.updated_dir = "templates\customer_updated.xlsx"
            self.create()
        else:
            self.processed_dir = "templates\팀 주간업무보고_V4.xlsx"
            make_dir('templates\Backup')
            shutil.copyfile(self.processed_dir, f'templates\Backup\팀 주간업무보고_V3_{datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.xlsx')
            self.updated_dir = "templates\팀 주간업무보고_V4.xlsx"
        self.template_file = None
        self.processed_file = None

        self.sheet = None

        self.week_range_text = f"업무진행현황 ({self.data_obj.search_on}:{str(self.week_start)[4:6]}/{str(self.week_start)[6:8]}-{str(self.week_end)[4:6]}/{str(self.week_end)[6:8]})"
        self.weekly_note_db = WeeklyNotesModel()

    def create(self):
        self.template_file = openpyxl.load_workbook(self.template_dir)
        for n, i in enumerate(self.customer_sequence):
            self.template_file.copy_worksheet(self.template_file.active).title = i
            self.sheet = self.template_file[i]
            self.sheet_condition(grid=False, zoom=85)
            self.fix_cell_size()
        self.template_file.remove(self.template_file['Sheet1'])
        self.template_file.save(self.processed_dir)

    def update(self):
        self.processed_file = openpyxl.load_workbook(self.processed_dir)
        if self.reload:
            self.reload_charts()

        customer_seq = []
        for n, i in enumerate(self.customer_sequence):
            try:
                self.sheet = self.processed_file[i]
            except KeyError:
                self.processed_file.copy_worksheet(self.sheet).title = i
                self.sheet = self.processed_file[i]
                print(f"고객사 {i}가 추가되었습니다.")
            self.delete_previous_images()
            self.delete_ranking_values()
            self.sheet_condition(grid=False, zoom=85)
            self.set_title(customer=i)
            self.attach_barplot(i, width=988, hight=580)
            self.attach_donutplot(i, width=270, hight=270)
            self.attach_delayplot(i, width=300, hight=270)
            self.update_ranking(i)
            self.fix_cell_size()
            self.set_title(customer=i)
            self.save_weeklynotes(customer=i)
            customer_seq.append(self.sheet)
        residual_sheets = [i for i in self.processed_file._sheets if i not in customer_seq]
        self.processed_file._sheets = residual_sheets + customer_seq
        self.select_sheet()
        self.processed_file.save(self.updated_dir)

    def delete_previous_images(self):
        self.sheet._images = []

    def update_ranking(self, customer):

        self.set_weekdates()

        for type_ in ['외관불량', '전체불량']:
            df = self.df_dict[f"{customer}_{type_}"]

            if df is None or len(df) == 0:
                continue
            elif type_ == '외관불량':
                for row in self.sheet['Y20:AC22']:
                    for col, cell in enumerate(row):
                        try:
                            cell.value = df.values[cell.row - 20][cell.column - 25]
                        except IndexError:
                            pass
                        if type(cell.value) == int:
                            cell.number_format = '#,###'
                        else:
                            cell.number_format = '#,##0.0'
            else:
                for row in self.sheet['Y23:AC25']:
                    for col, cell in enumerate(row):
                        try:
                            cell.value = df.values[cell.row-23][cell.column-25]
                        except IndexError:
                            pass
                        if type(cell.value) == int:
                            cell.number_format = '#,###'
                        else:
                            cell.number_format = '#,##0.0'

    def reload_charts(self):
        self.reload_barcharts()
        self.reload_donutcharts()
        self.reload_delayedcharts()

    def reload_barcharts(self):
        df_qty_this_year, df_qty_last_year_quarters, df_weekly, weekly_dates = self.data_obj.get_barplot_data()
        barplot_obj = CustomerBarplots(df_qty_this_year, df_qty_last_year_quarters, df_weekly, weekly_dates,
                                       regular_month=self.regular_month, yyyymm=self.yyyymm, weekly_customer=self.weekly_customer)
        barplot_obj.get_customer_barplot_set()
        self.customer_sequence = self.data_obj.customer_ranking_by_week(df_weekly, type_="외관_불량수량")

    def reload_donutcharts(self):
        ovs_df_year_type, ovs_df_month_type, types = self.data_obj.get_donutplot_data()
        donutplot_obj = CustomerDonutPlots(ovs_df_year_type, ovs_df_month_type, types, regular_month=self.regular_month,
                                           yyyymm=self.yyyymm, weekly_customer=self.weekly_customer)
        donutplot_obj.get_customer_donutplot_set()

    def reload_delayedcharts(self):
        df_acc = self.data_obj.get_delayed_days()
        delayedplot_obj = CustomerDelayedPlots(df_acc, regular_month=self.regular_month, yyyymm=self.yyyymm,
                                               weekly_customer=self.weekly_customer)
        delayedplot_obj.get_customer_delayed_plot_set()

    def sheet_condition(self, grid, zoom):
        self.sheet.sheet_view.showGridLines = grid
        self.sheet.sheet_view.zoomScale = zoom
        self.sheet.sheet_view.selection[0].activeCell = 'A1'
        self.sheet.sheet_view.selection[0].sqref = 'A1'
        self.sheet.sheet_view.topLeftCell = "A1"

    def set_title(self, customer):
        self.sheet['C2'] = f"■ {customer} 품질현황"
        self.sheet['AG2'] = f"■ 담당자 : {PERSON_IN_CHARGE[customer]}"

    def set_weekdates(self):
        self.sheet['AE19'] = self.week_range_text

    def save_weeklynotes(self, customer):
        this_week_task = self.sheet['W29'].value
        next_week_task = self.sheet['AD29'].value
        self.weekly_note_db.insert(customer=customer, poc=PERSON_IN_CHARGE[customer], week=self.week_range_text.split(":")[1][:-1],
                                   this_week_task=this_week_task, next_week_task=next_week_task)

    def barplot(self, customer):
        folder_name = os.path.join("spawn", "plots", datetime.now().strftime("%Y-%m-%d"))
        img_name = os.path.abspath(os.path.join(folder_name, f"{self.yyyymm}_{customer}_weekly.png"))
        return img_name

    def donutplot(self, customer):
        folder_name = os.path.join("spawn", "plots", datetime.now().strftime("%Y-%m-%d"))
        img_name_year = os.path.abspath(os.path.join(folder_name, f"{self.yyyymm}_{customer}_weekly_donut_year.png"))
        img_name_month = os.path.abspath(os.path.join(folder_name, f"{self.yyyymm}_{customer}_weekly_donut_month.png"))
        return img_name_year, img_name_month

    def delayplot(self, customer):
        folder_name = os.path.join("spawn", "plots", datetime.now().strftime("%Y-%m-%d"))
        img_name = os.path.abspath(os.path.join(folder_name, f"{self.yyyymm}_{customer}_weekly_delayed.png"))
        return img_name

    def attach_donutplot(self, customer, width, hight):
        img_year, img_month = self.donutplot(customer=customer)
        img_year = openpyxl.drawing.image.Image(img_year)
        img_year.anchor = 'Y5'
        img_year.height = hight
        img_year.width = width
        img_month = openpyxl.drawing.image.Image(img_month)
        img_month.anchor = 'AB5'
        img_month.height = hight
        img_month.width = width

        self.sheet.add_image(img_year)
        self.sheet.add_image(img_month)

    def attach_barplot(self, customer, width, hight):
        img = openpyxl.drawing.image.Image(self.barplot(customer=customer))
        img.anchor = 'D5'
        img.height = hight
        img.width = width
        self.sheet.add_image(img)

    def attach_delayplot(self, customer, width, hight):
        img = openpyxl.drawing.image.Image(self.delayplot(customer=customer))
        img.anchor = 'AF5'
        img.height = hight
        img.width = width
        self.sheet.add_image(img)

    def delete_ranking_values(self):
        for row in self.sheet['Y20:AC25']:  # 'Y20:AE25' --> ranking 유저기입부분 삭제
            for col, cell in enumerate(row):
                cell.value = ""

    def select_sheet(self):
        self.processed_file.active = 0

        for sheet in self.processed_file:
            if sheet.title == "주간업무보고":
                sheet.sheet_view.tabSelected = True
            else:
                sheet.sheet_view.tabSelected = False

    def fix_cell_size(self):
        self.sheet.column_dimensions['A'].width = 1.88
        self.sheet.column_dimensions['B'].width = 1.88
        self.sheet.column_dimensions['C'].width = 1.88

        for row in self.sheet['D1:V1']:
            for col, cell in enumerate(row):
                self.sheet.column_dimensions[get_column_letter(cell.column)].width = 6.89

        for row in self.sheet['W1:X1']:
            for col, cell in enumerate(row):
                self.sheet.column_dimensions[get_column_letter(cell.column)].width = 1.88

        self.sheet.column_dimensions['Y'].width = 4
        self.sheet.column_dimensions['Z'].width = 13
        self.sheet.column_dimensions['AA'].width = 20

        for row in self.sheet['AB1:AD1']:
            for col, cell in enumerate(row):
                self.sheet.column_dimensions[get_column_letter(cell.column)].width = 8.38

        self.sheet.column_dimensions['AE'].width = 9

        for row in self.sheet['AF1:AG1']:
            for col, cell in enumerate(row):
                self.sheet.column_dimensions[get_column_letter(cell.column)].width = 19.5

        self.sheet.column_dimensions['AH'].width = 1.88

        self.sheet.row_dimensions[1].height = 13.5
        self.sheet.row_dimensions[2].height = 29.25
        self.sheet.row_dimensions[3].height = 13.5

        for row in self.sheet['A4:A18']:
            for col, cell in enumerate(row):
                self.sheet.row_dimensions[cell.row].height = 16.40

        self.sheet.row_dimensions[19].height = 20

        for row in self.sheet['A20:A25']:
            for col, cell in enumerate(row):
                self.sheet.row_dimensions[cell.row].height = 33.5

        self.sheet.row_dimensions[26].height = 13.5
        self.sheet.row_dimensions[27].height = 20

        for row in self.sheet['A28:A40']:
            for col, cell in enumerate(row):
                self.sheet.row_dimensions[cell.row].height = 13.5


if __name__ == "__main__":
    import os
    os.chdir(os.pardir)
    os.chdir(os.pardir)

    excel = ExcelFormatSetter()
    # excel.create()
    excel.update()

    # os.startfile(excel.template_dir)
    # os.startfile(excel.processed_dir)
    os.startfile(excel.updated_dir)
