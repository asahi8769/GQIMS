# from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from utils_.config import CUSTOMERS, SMALL_CUSTOMERS
from tqdm import tqdm

# https://htmlcolorcodes.com/

brightpurple1 = PatternFill(start_color='E8DAEF',
                          end_color='E8DAEF',
                          fill_type='solid')

brightgreen1 = PatternFill(start_color='D5F5E3',
                          end_color='D5F5E3',
                          fill_type='solid')

brightgreen2 = PatternFill(start_color='ABEBC6',
                          end_color='ABEBC6',
                          fill_type='solid')

brightyellow = PatternFill(start_color='F9E79F',
                          end_color='F9E79F',
                          fill_type='solid')

brightorange = PatternFill(start_color='F5CBA7',
                          end_color='F5CBA7',
                          fill_type='solid')

brightpurple2 = PatternFill(start_color='D2B4DE',
                          end_color='D2B4DE',
                          fill_type='solid')

brightblue = PatternFill(start_color='AED6F1',
                          end_color='AED6F1',
                          fill_type='solid')

lightgrey = PatternFill(start_color='EAEDED',
                       end_color='EAEDED',
                       fill_type='solid')

darkgrey = PatternFill(start_color='AEB6BF',
                       end_color='AEB6BF',
                       fill_type='solid')

empasizing_yellow = PatternFill(start_color='FFF975',
                       end_color='FFF975',
                       fill_type='solid')


thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='566573'),
    right=Side(border_style=BORDER_THIN, color='566573'),
    top=Side(border_style=BORDER_THIN, color='566573'),
    bottom=Side(border_style=BORDER_THIN, color='566573')
)


def fit_column_width(sheet):

    column_widths = []
    for row in sheet:
        for i, cell in enumerate(row):
            try :
                number = int(cell.value)
            except (ValueError, TypeError):
                length = len(str(cell.value)) * 2 if cell.value is not None else 0
            else:
                length = len(str(cell.value)) * 1.4

            if len(column_widths) > i:
                if length > column_widths[i]:
                    column_widths[i] = length
            else:
                column_widths += [length]

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = int(column_width)


def get_col_from_index(col, add):
    return get_column_letter(column_index_from_string(col) + add)


def sheet_0_decorate(xfile, sheet_name_0, df_ovs_summary):

    sheet = xfile[sheet_name_0]
    fit_column_width(sheet)
    sheet.column_dimensions["A"].width = 2
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 100

    col_anchor = "B"
    row_anchor = 3
    col_len = len(df_ovs_summary.columns)
    index_depth = len(df_ovs_summary.index[0])
    df_len = len(df_ovs_summary)
    empasized_len = len(df_ovs_summary.loc[('종합', )])

    for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, index_depth-1)}{df_len+row_anchor}']:
        for col, cell in enumerate(row):
            cell.fill = brightyellow

    temp_anchor = get_col_from_index(col_anchor, index_depth)
    for row in sheet[f'{temp_anchor}{row_anchor}:{get_col_from_index(temp_anchor, col_len-1)}{df_len+row_anchor}']:
        for col, cell in enumerate(row):
            cell.border = thin_border
            if cell.value is None:
                cell.fill = lightgrey
            if cell.row == row_anchor:
                cell.fill = brightyellow
                sheet.column_dimensions[get_column_letter(cell.column)].width = 12
            if cell.value is not None and cell.value != 0 and cell.row != row_anchor:
                if type(cell.value) == int:
                    cell.number_format = '#,###'
                else:
                    cell.number_format = '#,##0.0'
            if row_anchor+1 <= cell.row <= row_anchor+empasized_len:
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)


def sheet_1_decorate(xfile, sheet_name_1, df_dict):
    sheet = xfile[sheet_name_1]
    fit_column_width(sheet)

    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 100

    row_n = 0
    row_anchor = 2
    for n, i in enumerate(df_dict.keys()):
        ticker = n % 2
        row = row_n * (len(df_dict[i]) + 1) + row_anchor
        if ticker == 1:
            row_n += 1
            row_anchor += 2
            sheet[f'{get_col_from_index("B", len(df_dict[i].columns) + 1)}{row}'] = '◎ ' + i
        else:
            sheet[f'B{row}'] = '◎ ' + i

    sheet.column_dimensions["A"].width = 2

    col_anchor = "B"
    row_anchor = 3
    col_len = len(df_dict[list(df_dict.keys())[0]].columns)
    for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, col_len * 2)}{row_anchor}']:
        for col, cell in enumerate(row):
            if cell.value in ['부품협력사', '품명']:
                sheet.column_dimensions[get_column_letter(cell.column)].width = 28
            if cell.value is None:
                sheet.column_dimensions[get_column_letter(cell.column)].width = 2

    for row in sheet:
        for col, cell in enumerate(row):
            if cell.value in df_dict[list(df_dict.keys())[0]].columns:
                cell.fill = brightyellow
            if cell.value is not None and not str(cell.value).startswith("◎"):
                cell.border = thin_border
            if str(cell.value).startswith("◎"):
                merge_end = get_col_from_index(coordinate_from_string(cell.coordinate)[0], col_len - 1)
                sheet.merge_cells(f'{cell.coordinate}:{merge_end}{coordinate_from_string(cell.coordinate)[1]}')
                cell.font = Font(size=12, bold=True)
                cell.alignment = Alignment(horizontal='center')
            if cell.value is not None and cell.value != 0:
                if type(cell.value) == int:
                    cell.number_format = '#,###'
                else:
                    cell.number_format = '#,##0.0'


def sheet_2_decorate(xfile, sheet_name_2, ovs_df, ovs_download_date, ovs_now):
    sheet = xfile[sheet_name_2]
    fit_column_width(sheet)
    sheet.column_dimensions["A"].width = 2
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 95
    sheet.freeze_panes = 'B5'

    col_anchor = "B"
    row_anchor = 4
    ovs_len = len(ovs_df)
    ovs_col_len = len(ovs_df.columns)
    ovs_idx_len = len(ovs_df.index[0])

    sheet[f'{col_anchor}1'] = "다운로드일자"
    sheet[f'{get_col_from_index(col_anchor, 1)}1'] = ovs_download_date
    sheet[f'{col_anchor}2'] = "자료작성일자"
    sheet[f'{get_col_from_index(col_anchor, 1)}2'] = ovs_now[:10]

    for row in sheet[f'{col_anchor}1:{col_anchor}2']:
        for col, cell in enumerate(row):
            row[col].fill = brightgreen1

    for row in sheet[f'{col_anchor}1:{get_col_from_index(col_anchor, 1)}2']:
        for col, cell in enumerate(row):
            cell.border = thin_border

    customers_sections = 35 * len(CUSTOMERS) + row_anchor
    for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, ovs_idx_len-1)}{customers_sections}']:
        for col, cell in enumerate(row):
            cell.fill = brightyellow

    for row in sheet[f'{col_anchor}{customers_sections+1}:{get_col_from_index(col_anchor, ovs_idx_len-1)}{customers_sections+6}']:
        for col, cell in enumerate(row):
            cell.fill = brightgreen1

    for row in sheet[f'{col_anchor}{customers_sections+7}:{get_col_from_index(col_anchor, ovs_idx_len-1)}{customers_sections+44}']:
        for col, cell in enumerate(row):
            cell.fill = brightgreen2

    for row in sheet[f'{col_anchor}{customers_sections+45}:{get_col_from_index(col_anchor, ovs_idx_len-1)}{customers_sections+66}']:
        for col, cell in enumerate(row):
            cell.fill = brightblue

    for row in sheet[f'{col_anchor}{customers_sections+67}:{get_col_from_index(col_anchor, ovs_idx_len-1)}{customers_sections+71}']:
        for col, cell in enumerate(row):
            cell.fill = brightpurple2

    for row in sheet[f'{col_anchor}{customers_sections+72}:{get_col_from_index(col_anchor, ovs_idx_len-1)}{(customers_sections+71)+(10*len(SMALL_CUSTOMERS))}']:
        for col, cell in enumerate(row):
            cell.fill = brightorange

    for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, ovs_idx_len+ovs_col_len-1)}{row_anchor}']:
        for col, cell in enumerate(row):
            cell.fill = brightyellow
            if cell.value in ['구분1', '구분2', '구분3']:
                sheet.column_dimensions[get_column_letter(cell.column)].width = 16
            if cell.value == '구분4':
                sheet.column_dimensions[get_column_letter(cell.column)].width = 8

    table_anchor = f"{get_col_from_index(col_anchor, ovs_idx_len)}"

    for row in sheet[f'{table_anchor}{row_anchor+1}:{get_col_from_index(table_anchor, ovs_col_len-1)}{row_anchor+ovs_len}']:
        for col, cell in enumerate(row):
            cell.border = thin_border
            if cell.value is None:
                cell.fill = lightgrey
            if cell.value is not None and cell.value != 0:
                if type(cell.value) == int:
                    cell.number_format = '#,###'
                else:
                    cell.number_format = '#,##0.0'

    for n, customer in enumerate(CUSTOMERS):
        emphasized_len = len(ovs_df.loc[('HAOS',)])
        row_gqms = 5 + emphasized_len*n
        sqd_gqims = 11 + emphasized_len * n
        apq_gqims = 25 + emphasized_len * n

        for row in sheet[f'{table_anchor}{row_gqms}:{get_col_from_index(table_anchor, ovs_col_len-1)}{row_gqms+1}']:
            for col, cell in enumerate(row):
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)

        for row in sheet[f'{table_anchor}{sqd_gqims}:{get_col_from_index(table_anchor, ovs_col_len-1)}{sqd_gqims+1}']:
            for col, cell in enumerate(row):
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)

        for row in sheet[f'{table_anchor}{apq_gqims}:{get_col_from_index(table_anchor, ovs_col_len-1)}{apq_gqims+1}']:
            for col, cell in enumerate(row):
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)

    for n, customer in enumerate(SMALL_CUSTOMERS):
        emphasized_len = len(ovs_df.loc[('HMI',)])
        sqd_gqims = (customers_sections+72) + emphasized_len * n

        for row in sheet[f'{table_anchor}{sqd_gqims}:{get_col_from_index(table_anchor, ovs_col_len-1)}{sqd_gqims+1}']:
            for col, cell in enumerate(row):
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)

    for row in sheet[f'{table_anchor}{customers_sections+1}:{get_col_from_index(table_anchor, ovs_col_len-1)}{customers_sections+2}']:
        for col, cell in enumerate(row):
            cell.fill = empasizing_yellow
            cell.font = Font(bold=True)

    for row in sheet[f'{table_anchor}{customers_sections+7}:{get_col_from_index(table_anchor, ovs_col_len-1)}{customers_sections+8}']:
        for col, cell in enumerate(row):
            cell.fill = empasizing_yellow
            cell.font = Font(bold=True)

    for row in sheet[f'{table_anchor}{customers_sections+21}:{get_col_from_index(table_anchor, ovs_col_len-1)}{customers_sections+22}']:
        for col, cell in enumerate(row):
            cell.fill = empasizing_yellow
            cell.font = Font(bold=True)

    for row in sheet[f'{table_anchor}{customers_sections+33}:{get_col_from_index(table_anchor, ovs_col_len-1)}{customers_sections+34}']:
        for col, cell in enumerate(row):
            cell.fill = empasizing_yellow
            cell.font = Font(bold=True)

    for row in sheet[f'{table_anchor}{customers_sections+45}:{get_col_from_index(table_anchor, ovs_col_len-1)}{customers_sections+47}']:
        for col, cell in enumerate(row):
            cell.fill = empasizing_yellow
            cell.font = Font(bold=True)

    for row in sheet[f'{table_anchor}{customers_sections+67}:{get_col_from_index(table_anchor, ovs_col_len-1)}{customers_sections+71}']:
        for col, cell in enumerate(row):
            cell.fill = brightpurple1
            cell.font = Font(bold=True)


def sheet_3_decorate(xfile, sheet_name_3, dms_df, dms_summary, dms_summary_regional, dms_download_date, dms_now):
    sheet = xfile[sheet_name_3]
    fit_column_width(sheet)
    sheet.column_dimensions["A"].width = 2

    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85

    col_anchor = "B"
    sheet[f'{col_anchor}1'] = "다운로드일자"
    sheet[f'{get_col_from_index(col_anchor, 1)}1'] = dms_download_date
    sheet[f'{col_anchor}2'] = "자료작성일자"
    sheet[f'{get_col_from_index(col_anchor, 1)}2'] = dms_now[:10]

    for row in sheet[f'{col_anchor}1:{col_anchor}2']:
        for col, cell in enumerate(row):
            cell.fill = brightgreen1

    for row in sheet[f'{col_anchor}1:{get_col_from_index(col_anchor, 1)}2']:
        for col, cell in enumerate(row):
            cell.border = thin_border

    row_anchor = 4
    col_len = len(dms_df.columns)
    index_depth = len(dms_df.index[0])
    df_len = len(dms_df)
    emphasized_len = len(dms_df.loc[('국내입고불량 실적',)])
    sheet.column_dimensions[get_col_from_index(col_anchor, col_len + 2)].width = 2

    for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, col_len + 1)}{row_anchor}']:
        for col, cell in enumerate(row):
            if cell.value == '구분2':
                sheet.column_dimensions[get_column_letter(cell.column)].width = 11

    for row in sheet[
        f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, index_depth - 1)}{row_anchor + df_len}']:
        for col, cell in enumerate(row):
            cell.fill = brightgreen2

    temp_anchor = get_col_from_index(col_anchor, index_depth)
    col_len = len(dms_df.columns)
    for row in sheet[f'{temp_anchor}{row_anchor}:{get_col_from_index(temp_anchor, col_len - 1)}{row_anchor}']:
        for col, cell in enumerate(row):
            cell.fill = brightgreen2

    ppm_row = [1, 4, 7]
    for row in sheet[
        f'{temp_anchor}{row_anchor + 1}:{get_col_from_index(temp_anchor, col_len - 1)}{row_anchor + df_len}']:
        for col, cell in enumerate(row):
            cell.border = thin_border
            if cell.value is None:
                cell.fill = lightgrey
            if row_anchor + 1 <= cell.row <= row_anchor + emphasized_len:
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)
            if cell.row in [row_anchor + emphasized_len*3 + i for i in ppm_row]:
                cell.fill = empasizing_yellow
                cell.font = Font(bold=True)

    sum_col_len = len(dms_summary.columns)
    sum_len = len(dms_summary)
    sum_header_anchor = f'{get_col_from_index(temp_anchor, col_len + 1)}'
    for row in sheet[
        f'{sum_header_anchor}{row_anchor}:{get_col_from_index(sum_header_anchor, sum_col_len - 1)}{row_anchor + sum_len}']:
        for col, cell in enumerate(row):
            cell.border = thin_border
            if cell.row == row_anchor:
                cell.fill = brightgreen2

    region_len = len(dms_summary_regional)
    region_col_len = len(dms_summary_regional.columns)
    for row in sheet[
        f'{sum_header_anchor}{row_anchor + sum_len + 3}:{get_col_from_index(sum_header_anchor, region_col_len - 1)}{row_anchor + sum_len + 3 + region_len}']:
        for col, cell in enumerate(row):
            cell.border = thin_border
            if cell.row == row_anchor + sum_len + 3:
                cell.fill = brightgreen2

    for row in sheet:
        for col, cell in enumerate(row):
            if cell.value is not None and cell.value != 0 and cell.row != row_anchor:
                if type(cell.value) == int:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.0'


def graph_sheet_decorate(xfile, graph_sheet, plot_name_1, plot_name_2, plot_name_3):
    sheet = xfile.create_sheet(graph_sheet, 0)
    sheet.column_dimensions["A"].width = 2
    sheet.column_dimensions["O"].width = 2
    sheet.sheet_view.showGridLines = False

    img_1 = openpyxl.drawing.image.Image(plot_name_1)
    img_1.anchor = 'B2'
    img_2 = openpyxl.drawing.image.Image(plot_name_2)
    img_2.anchor = 'B45'
    img_3 = openpyxl.drawing.image.Image(plot_name_3)
    img_3.anchor = 'P45'
    sheet.add_image(img_1)
    sheet.add_image(img_2)
    sheet.add_image(img_3)
    sheet.sheet_view.zoomScale = 95


def raw_sheet_decorate(xfile, sheet_name, df):
    sheet = xfile[sheet_name]
    fit_column_width(sheet)

    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85

    header = df.columns.tolist()

    row_anchor = 2
    col_anchor = "A"
    region_len = len(df)
    region_col_len = len(df.columns)

    for row in sheet[f'{col_anchor}{row_anchor}:{get_col_from_index(col_anchor, region_col_len - 1)}{row_anchor + region_len}']:
        for col, cell in enumerate(row):
            cell.border = thin_border
            if cell.value in header:
                cell.fill = brightyellow
                if cell.value in ['차종', '코드', '코드.1', '작성자', '담당자', '현재결재자', '사업장']:
                    sheet.column_dimensions[get_column_letter(cell.column)].width = 8
                if cell.value in ['품번', '발생일자', '등록일자', '해외결재일', '국내접수일', '국내결재일', '제기일자', '제기부서',
                                  '검사일자', '통보일자', '접수일자', '결재일자', '포장장' ,'대책등록일', '대책확인일', '개선일',
                                  '불량유형1', '불량유형2', '불량유형3', 'P/T기종', 'P/T_Model', '고객사', 'KDQR_Type',
                                  'PPR_확인일자', '발생원인(대)', '발생원인(중)', '발생원인(소)', '대책등록일', '대책확인일',
                                  '대책확인일(해외)', '개선일', '공장', '불량원인(대)', '불량원인(중)', '불량원인(소)', '귀책구분',
                                  '불량유형(대)', '불량유형(중)', '불량유형(소)', 'KDQR_No', 'PPR_확인_일자']:
                    sheet.column_dimensions[get_column_letter(cell.column)].width = 18
                if cell.value in ['부품협력사', '귀책처', '제목', '품명', '최종조립처', '긴급도']:
                    sheet.column_dimensions[get_column_letter(cell.column)].width = 35


def sheet_decorate(file, sheet_name_0, sheet_name_1, sheet_name_2, sheet_name_3, graph_sheet, df_ovs_summary, ovs_df,
                   ovs_download_date, ovs_now, dms_download_date, dms_now, plot_name_1, df_dict, dms_df, dms_summary,
                   dms_summary_regional, sheet_name_4, sheet_name_5, sheet_name_6, sheet_name_7, df_ppr, df_ovs, df_dms,
                   df_ihr, plot_name_2, plot_name_3):

    xfile = openpyxl.load_workbook(file)

    sheet_0_decorate(xfile, sheet_name_0, df_ovs_summary)
    sheet_1_decorate(xfile, sheet_name_1, df_dict)
    sheet_2_decorate(xfile, sheet_name_2, ovs_df, ovs_download_date, ovs_now)
    sheet_3_decorate(xfile, sheet_name_3, dms_df, dms_summary, dms_summary_regional, dms_download_date, dms_now)
    graph_sheet_decorate(xfile, graph_sheet, plot_name_1, plot_name_2, plot_name_3)

    for raw_sheet, df in tqdm(
            list(zip([sheet_name_4, sheet_name_5, sheet_name_6, sheet_name_7], [df_ppr, df_ovs, df_dms, df_ihr]))):
        raw_sheet_decorate(xfile, raw_sheet, df)

    for sheet in xfile:
        if sheet.title == graph_sheet:
            sheet.sheet_view.tabSelected = True
        else:
            sheet.sheet_view.tabSelected = False

    xfile.save(file)


